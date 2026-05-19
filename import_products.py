"""
import_products.py — ToolPrime
Импорт товаров из Excel-прайса в базу данных SQLite через SQLAlchemy.

Структура прайса (Лист1):
  Модификация | Название | Ед.изм. | URL | Торговая марка | Название группы | Название подгруппы | Цена

Запуск:
  python import_products.py                          # прайс рядом со скриптом
  python import_products.py --file path/to/file.xlsx
  python import_products.py --clear                  # очистить таблицу перед импортом
"""

import argparse
import re
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

# ── Настройки ─────────────────────────────────────────────────────────────────

DEFAULT_EXCEL = Path(__file__).parent / "ПРАЙС_25_02_26__скидка_20_.xlsx"
DATABASE_URL  = f"sqlite:///{Path(__file__).parent / 'toolprime.db'}"
SHEET_NAME    = "Лист1"          # лист с данными
BATCH_SIZE    = 500              # сколько товаров вставлять за одну транзакцию
DISCOUNT      = 0.20             # скидка 20% уже заложена в прайсе; поле для отображения

# ── Импорт моделей ─────────────────────────────────────────────────────────────
# Предполагается, что модели описаны в models.py рядом со скриптом.
# Если у вас другой модуль — исправьте импорт ниже.
try:
    from models import Base, Product, Category, Brand
except ImportError:
    sys.exit(
        "❌  Не найден models.py. Убедитесь, что файл находится рядом с import_products.py.\n"
        "    Ожидаемые классы: Base, Product, Category, Brand."
    )

# ── Вспомогательные функции ────────────────────────────────────────────────────

def slugify(text: str) -> str:
    """Транслитерация + приведение к slug-формату."""
    ru_to_en = {
        'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'zh',
        'з':'z','и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o',
        'п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'kh','ц':'ts',
        'ч':'ch','ш':'sh','щ':'shch','ъ':'','ы':'y','ь':'','э':'e','ю':'yu',
        'я':'ya',
    }
    result = text.lower()
    result = ''.join(ru_to_en.get(c, c) for c in result)
    result = re.sub(r'[^a-z0-9]+', '-', result)
    result = result.strip('-')
    return result[:120]  # ограничение длины slug


def make_unique_slug(base_slug: str, existing: set[str]) -> str:
    """Добавляет числовой суффикс, если slug уже занят."""
    slug = base_slug
    counter = 1
    while slug in existing:
        slug = f"{base_slug}-{counter}"
        counter += 1
    existing.add(slug)
    return slug


def clean_category_name(raw: str) -> str:
    """Убирает числовой префикс вида '02_' из названия категории."""
    return re.sub(r'^\d+_', '', raw).strip()


def progress(current: int, total: int, start_time: float, label: str = "") -> None:
    """Однострочный прогресс-бар в терминале."""
    pct   = current / total * 100
    done  = int(pct / 2)
    bar   = "█" * done + "░" * (50 - done)
    elapsed = time.time() - start_time
    speed   = current / elapsed if elapsed > 0 else 0
    eta     = (total - current) / speed if speed > 0 else 0
    print(
        f"\r  [{bar}] {pct:5.1f}%  {current:>6}/{total}"
        f"  {speed:>6.0f} т/с  ETA {eta:>4.0f}с  {label}",
        end="", flush=True,
    )


# ── Основная логика ────────────────────────────────────────────────────────────

def run(excel_path: Path, clear: bool) -> None:
    print(f"\n{'='*60}")
    print(f"  ToolPrime — импорт товаров")
    print(f"  Файл:  {excel_path.name}")
    print(f"  БД:    {DATABASE_URL}")
    print(f"{'='*60}\n")

    # 1. Подключение к БД
    engine = create_engine(DATABASE_URL, echo=False)
    Base.metadata.create_all(engine)          # создаём таблицы, если их нет

    with Session(engine) as session:

        # 2. Опциональная очистка
        if clear:
            print("  ⚠️  --clear: удаляем все товары, категории, бренды...")
            session.execute(text("DELETE FROM products"))
            session.execute(text("DELETE FROM categories"))
            session.execute(text("DELETE FROM brands"))
            session.commit()
            print("  ✅  Таблицы очищены.\n")

        # 3. Загрузка существующих данных для дедупликации
        existing_skus   = {p.sku for p in session.query(Product.sku)}
        existing_slugs  = {p.slug for p in session.query(Product.slug)}
        cat_cache: dict[str, Category]  = {}   # slug → Category
        brand_cache: dict[str, Brand]   = {}   # slug → Brand

        # Подгрузим уже имеющиеся категории и бренды
        for c in session.query(Category).all():
            cat_cache[c.slug] = c
        for b in session.query(Brand).all():
            brand_cache[b.slug] = b

        # 4. Открываем Excel
        print("  📂  Открываем Excel...")
        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as e:
            sys.exit(f"❌  Не удалось открыть файл: {e}")

        if SHEET_NAME not in wb.sheetnames:
            sys.exit(f"❌  Лист '{SHEET_NAME}' не найден. Доступные: {wb.sheetnames}")

        ws = wb[SHEET_NAME]

        # Считаем строки (пропускаем заголовок)
        print("  🔢  Считаем строки...")
        total_rows = sum(
            1 for row in ws.iter_rows(min_row=2, values_only=True)
            if any(c is not None for c in row)
        )
        print(f"  📊  Найдено строк: {total_rows}\n")

        # 5. Импорт
        added = skipped_dup = skipped_no_price = 0
        batch: list[Product] = []
        start = time.time()

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            # Пустая строка — пропускаем
            if not any(c is not None for c in row):
                continue

            sku, name, unit, url, brand_name, cat_name, subcat_name, price = row

            # ── Валидация ──────────────────────────────────────────────────
            sku        = str(sku).strip()  if sku        else None
            name       = str(name).strip() if name       else None
            brand_name = str(brand_name).strip() if brand_name else "Без бренда"
            cat_name   = str(cat_name).strip()   if cat_name   else "Прочее"
            subcat_name= str(subcat_name).strip() if subcat_name else None
            unit       = str(unit).strip() if unit else "шт"
            url        = str(url).strip()  if url  else None

            if not name or not sku:
                skipped_dup += 1
                continue

            if price is None:
                skipped_no_price += 1
                continue

            try:
                price = float(price)
            except (ValueError, TypeError):
                skipped_no_price += 1
                continue

            # ── Дедупликация по SKU ────────────────────────────────────────
            if sku in existing_skus:
                skipped_dup += 1
                continue
            existing_skus.add(sku)

            # ── Категория (основная) ───────────────────────────────────────
            clean_cat = clean_category_name(cat_name)
            cat_slug  = slugify(clean_cat)
            if cat_slug not in cat_cache:
                cat = Category(
                    name=clean_cat,
                    slug=cat_slug,
                    parent_id=None,
                )
                session.add(cat)
                session.flush()           # получаем id до коммита
                cat_cache[cat_slug] = cat
            category = cat_cache[cat_slug]

            # ── Подкатегория (дочерняя) ────────────────────────────────────
            subcat = None
            if subcat_name:
                clean_sub  = clean_category_name(subcat_name)
                sub_slug   = slugify(f"{cat_slug}-{clean_sub}")
                if sub_slug not in cat_cache:
                    subcat = Category(
                        name=clean_sub,
                        slug=sub_slug,
                        parent_id=category.id,
                    )
                    session.add(subcat)
                    session.flush()
                    cat_cache[sub_slug] = subcat
                subcat = cat_cache[sub_slug]

            # ── Бренд ──────────────────────────────────────────────────────
            brand_slug = slugify(brand_name)
            if brand_slug not in brand_cache:
                brand = Brand(name=brand_name, slug=brand_slug)
                session.add(brand)
                session.flush()
                brand_cache[brand_slug] = brand
            brand = brand_cache[brand_slug]

            # ── Slug товара ────────────────────────────────────────────────
            base_slug    = slugify(f"{name}-{sku}")
            product_slug = make_unique_slug(base_slug, existing_slugs)

            # ── Создание товара ────────────────────────────────────────────
            product = Product(
                sku          = sku,
                name         = name,
                slug         = product_slug,
                price        = round(price, 2),
                unit         = unit,
                brand_id     = brand.id,
                category_id  = (subcat or category).id,
                in_stock     = True,
                description  = f"Артикул: {sku}. {name}.",
                image_url    = None,
                source_url   = url,
            )
            batch.append(product)
            added += 1

            # ── Батчевая вставка ───────────────────────────────────────────
            if len(batch) >= BATCH_SIZE:
                session.bulk_save_objects(batch)
                session.commit()
                batch.clear()

            # ── Прогресс ───────────────────────────────────────────────────
            if idx % 100 == 0 or idx == total_rows:
                progress(idx, total_rows, start)

        # Вставляем остаток
        if batch:
            session.bulk_save_objects(batch)
            session.commit()

        print()  # перенос после прогресс-бара

    # 6. Итог
    elapsed = time.time() - start
    print(f"\n{'='*60}")
    print(f"  ✅  Импорт завершён за {elapsed:.1f} с")
    print(f"  ➕  Добавлено товаров:      {added:>6}")
    print(f"  ⏭   Пропущено (дубли):      {skipped_dup:>6}")
    print(f"  ⚠️   Пропущено (нет цены):  {skipped_no_price:>6}")
    print(f"  🏷   Брендов в БД:           {len(brand_cache):>6}")
    print(f"  📁  Категорий в БД:          {len(cat_cache):>6}")
    print(f"{'='*60}\n")


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Импорт прайса ToolPrime в SQLite")
    parser.add_argument(
        "--file", type=Path, default=DEFAULT_EXCEL,
        help="Путь к Excel-файлу (по умолчанию: прайс рядом со скриптом)",
    )
    parser.add_argument(
        "--clear", action="store_true",
        help="Очистить товары/категории/бренды перед импортом",
    )
    args = parser.parse_args()

    if not args.file.exists():
        sys.exit(f"❌  Файл не найден: {args.file}")

    run(args.file, args.clear)
